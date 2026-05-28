# ============================================================
#  Neminath Wood Industry Pvt. Ltd.
#  Attendance System — utils.py
#  App: app
# ============================================================

import calendar
import datetime
from decimal import Decimal
from django.utils import timezone


# ------------------------------------------------------------
# WORKING DAY CALCULATOR
# ------------------------------------------------------------

# def is_second_or_fourth_saturday(date):
#     """Returns True if date is 2nd or 4th Saturday of its month."""
#     if date.weekday() != 5:  # 5 = Saturday
#         return False
#     saturday_number = (date.day - 1) // 7 + 1
#     return saturday_number in (2, 4)


def is_working_day(date):
    """
    Returns True if the date is a working day.
    Holidays:
      - Every Sunday
      - 2nd and 4th Saturday
      - Any date in Holiday table
    """
    from .models import Holiday

    # Sunday
    if date.weekday() == 6:
        return False

    # # 2nd or 4th Saturday
    # if is_second_or_fourth_saturday(date):
    #     return False

    # Company holidays
    if Holiday.objects.filter(date=date).exists():
        return False

    return True


def get_working_days_list(year, month):
    """Returns list of working day dates in a given month."""
    total_days   = calendar.monthrange(year, month)[1]
    working_days = []
    for day in range(1, total_days + 1):
        date = datetime.date(year, month, day)
        if is_working_day(date):
            working_days.append(date)
    return working_days


def get_working_days_count(year, month):
    """Returns count of working days in a given month."""
    return len(get_working_days_list(year, month))


def get_month_name(month):
    return calendar.month_name[month]


# ------------------------------------------------------------
# TIME HELPERS
# ------------------------------------------------------------

IN_GRACE_END    = datetime.time(9,  0)   # Late after 9:00 AM
OUT_GRACE_START = datetime.time(18, 50)  # Early exit before 6:50 PM
WORK_MINUTES_PER_DAY   = 610             # 10h 10min
MONTHLY_BUFFER_MINUTES = 120             # 2 hours free per month


def compute_late_minutes(in_time, date=None):
    """Returns minutes late. 0 if on time."""
    if not in_time or in_time <= IN_GRACE_END:
        return 0
    d        = date or datetime.date.today()
    in_dt    = datetime.datetime.combine(d, in_time)
    grace_dt = datetime.datetime.combine(d, IN_GRACE_END)
    return max(int((in_dt - grace_dt).total_seconds() / 60), 0)


def compute_early_exit_minutes(out_time, date=None):
    """Returns minutes of early exit. 0 if on time or invalid."""
    if not out_time:
        return 0
    # Guard — out time must be afternoon
    if out_time <= datetime.time(12, 0):
        return 0
    if out_time >= OUT_GRACE_START:
        return 0
    d        = date or datetime.date.today()
    out_dt   = datetime.datetime.combine(d, out_time)
    grace_dt = datetime.datetime.combine(d, OUT_GRACE_START)
    return max(int((grace_dt - out_dt).total_seconds() / 60), 0)


# ------------------------------------------------------------
# PERMISSION BUDGET
# ------------------------------------------------------------

def get_permission_budget(employee, year, month):
    """
    Returns how much permission the employee has used this month.
    {used, remaining, limit, exceeded, percent}
    """
    from .models import Permission
    from django.db.models import Sum

    used = Permission.objects.filter(
        employee    = employee,
        date__year  = year,
        date__month = month,
        status      = 'approved',
    ).aggregate(total=Sum('requested_minutes'))['total'] or 0

    remaining = max(MONTHLY_BUFFER_MINUTES - used, 0)
    percent   = min(int((used / MONTHLY_BUFFER_MINUTES) * 100), 100)

    return {
        'limit':     MONTHLY_BUFFER_MINUTES,
        'used':      used,
        'remaining': remaining,
        'exceeded':  used >= MONTHLY_BUFFER_MINUTES,
        'percent':   percent,
    }


# ------------------------------------------------------------
# SALARY CALCULATOR
# ------------------------------------------------------------

def calculate_net_salary(employee, year, month):
    """
    Full salary calculation for one employee for one month.
    Returns a dict with all breakdown values.
    """
    from .models import AttendanceRecord, Permission
    from django.db.models import Sum

    working_days = get_working_days_count(year, month)

    if working_days == 0:
        return {}

    records = AttendanceRecord.objects.filter(
        employee    = employee,
        date__year  = year,
        date__month = month,
    )

    present_days    = records.filter(status='present').count()
    # Count ALL present records (including holidays worked)
    # present_days = records.filter(status='present').count()
    paid_leave_days = records.filter(status='paid_leave').count()
    absent_days     = max(working_days - present_days - paid_leave_days, 0)

    agg = records.filter(status='present').aggregate(
        late  = Sum('late_minutes'),
        early = Sum('early_exit_minutes'),
    )
    total_late       = agg['late']  or 0
    total_early_exit = agg['early'] or 0

    total_permissions = Permission.objects.filter(
        employee    = employee,
        date__year  = year,
        date__month = month,
        status      = 'approved',
    ).aggregate(total=Sum('requested_minutes'))['total'] or 0

    total_deviation = total_late + total_early_exit + total_permissions
    excess_minutes  = max(total_deviation - MONTHLY_BUFFER_MINUTES, 0)

    monthly_salary   = employee.monthly_salary
    # per_day          = monthly_salary / Decimal(working_days)
    import calendar
    calendar_days = calendar.monthrange(year, month)[1]
    per_day = monthly_salary / Decimal(calendar_days)
    absent_deduction = (per_day * Decimal(absent_days)).quantize(Decimal('0.01'))

    total_work_mins = Decimal(working_days * WORK_MINUTES_PER_DAY)
    auto_debit      = (
        Decimal(excess_minutes) / total_work_mins * monthly_salary
    ).quantize(Decimal('0.01')) if excess_minutes > 0 and total_work_mins > 0 else Decimal('0.00')

    return {
        'employee':                 employee,
        'year':                     year,
        'month':                    month,
        'month_name':               get_month_name(month),
        'total_working_days':       working_days,
        'present_days':             present_days,
        'paid_leave_days':          paid_leave_days,
        'absent_days':              absent_days,
        'total_late_minutes':       total_late,
        'total_early_exit_minutes': total_early_exit,
        'total_permission_minutes': total_permissions,
        'total_deviation_minutes':  total_deviation,
        'excess_minutes':           excess_minutes,
        'monthly_salary':           monthly_salary,
        'absent_deduction':         absent_deduction,
        'auto_debit_amount':        auto_debit,
    }


# ------------------------------------------------------------
# DASHBOARD SUMMARY
# ------------------------------------------------------------

def get_today_summary():
    """Returns today's attendance counts."""
    from .models import AttendanceRecord, Employee
    today        = timezone.localdate()
    total_active = Employee.objects.filter(status='active').count()
    records      = AttendanceRecord.objects.filter(date=today)
    present      = records.filter(status='present').count()
    absent       = records.filter(status='absent').count()
    late         = records.filter(status='present', late_minutes__gt=0).count()
    paid_leave   = records.filter(status='paid_leave').count()

    return {
        'today':        today,
        'total_active': total_active,
        'present':      present,
        'absent':       absent,
        'late':         late,
        'paid_leave':   paid_leave,
        'not_marked':   total_active - present - absent - paid_leave,
        'is_working':   is_working_day(today),
    }


# ------------------------------------------------------------
# BULK ATTENDANCE HELPER
# ------------------------------------------------------------

def get_or_create_daily_attendance(date):
    """
    Returns queryset of all active employees with their attendance
    record for the given date. Creates absent records if missing.
    Used in daily attendance entry page.
    """
    from .models import Employee, AttendanceRecord

    employees = Employee.objects.filter(status='active').select_related('department')

    if is_working_day(date):
        for emp in employees:
            AttendanceRecord.objects.get_or_create(
                employee = emp,
                date     = date,
                defaults = {'status': 'absent'},
            )

    records = AttendanceRecord.objects.filter(
        date     = date,
        employee__status = 'active',
    ).select_related('employee', 'employee__department').order_by('employee__employee_id')

    return records


# ============================================================
#  Neminath Wood Industry Pvt. Ltd.
#  export_utils.py — Excel + PDF generators from real DB
# ============================================================

import os
import io
import calendar
import datetime
from decimal import Decimal

# ── Excel ──
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# ── PDF ──
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle,
    Paragraph, Spacer, Image as RLImage, HRFlowable,
)
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# ── Django ──
from django.conf import settings


# ── Logo path helper ──
def _logo_path():
    candidates = [
        os.path.join(settings.BASE_DIR, 'app',    'static', 'app',    'NWIPLogo.jpeg'),
        os.path.join(settings.BASE_DIR, 'static', 'app',    'NWIPLogo.jpeg'),
        os.path.join(settings.BASE_DIR, 'app',    'static', 'NWIPLogo.jpeg'),
        os.path.join(settings.BASE_DIR, 'NWIPLogo.jpeg'),
    ]
    for p in candidates:
        if os.path.exists(p):
            return p
    return None


# ============================================================
# EXCEL — MONTHLY SALARY REPORT
# ============================================================

def generate_monthly_excel(report_data, totals, year, month, working_days):
    """
    Generates styled .xlsx for the monthly salary report.
    report_data : list of dicts from monthly_report view
    Returns     : BytesIO buffer
    """
    month_name = calendar.month_name[month]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{month_name[:3]} {year}"

    # ── Palette ──
    C_HDR_BG   = "1A3A2A"
    C_HDR_FG   = "F5C842"
    C_COL_BG   = "2D5A3D"
    C_COL_FG   = "FFFFFF"
    C_PRESENT  = "D6F5E3"
    C_ABSENT   = "FFE0E0"
    C_LATE     = "FFF3CD"
    C_EXCESS   = "FFD0B0"
    C_NET_BG   = "1A3A2A"
    C_NET_FG   = "F5C842"
    C_TOT_BG   = "2D5A3D"
    C_TOT_FG   = "FFFFFF"
    C_ALT1     = "F4F8F5"
    C_BORDER   = "2D5A3D"

    def bdr(color=C_BORDER, s="thin"):
        sd = Side(style=s, color=color)
        return Border(left=sd, right=sd, top=sd, bottom=sd)

    # ── Row heights ──
    ws.row_dimensions[1].height = 72
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 8

    # ── Logo ──
    logo = _logo_path()
    if logo:
        try:
            img = XLImage(logo)
            img.height, img.width = 68, 68
            img.anchor = "A1"
            ws.add_image(img)
        except Exception:
            pass

    # ── Header ──
    ws.merge_cells("B1:Q1")
    c = ws["B1"]
    c.value     = "Neminath Wood Industry Pvt. Ltd."
    c.font      = Font(name="Arial", size=18, bold=True, color=C_HDR_FG)
    c.fill      = PatternFill("solid", fgColor=C_HDR_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("B2:Q2")
    c = ws["B2"]
    c.value     = f"MONTHLY SALARY REPORT — {month_name.upper()} {year}"
    c.font      = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor=C_HDR_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("B3:Q3")
    c = ws["B3"]
    c.value     = f"Working Days: {working_days}   |   Generated: {datetime.date.today().strftime('%d %B %Y')}   |   Confidential"
    c.font      = Font(name="Arial", size=9, italic=True, color="AAAAAA")
    c.fill      = PatternFill("solid", fgColor=C_HDR_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")

    for r in [1, 2, 3, 4]:
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=C_HDR_BG)

    # ── Column headers ──
    headers = [
        ("S.No",           4),  ("Emp ID",       10), ("Employee Name",  22),
        ("Department",    15),  ("Designation",  15), ("Present",         8),
        ("Absent",         8),  ("Paid Leave",    9), ("Late (min)",       9),
        ("Early (min)",    9),  ("Perm (min)",    9), ("Excess (min)",     9),
        ("Gross (Rs.)",   13),  ("-Absent (Rs.)",12), ("Auto Debit (Rs.)",13),
        ("Override (Rs.)",13),  ("Net Salary",   14),
    ]

    HDR_ROW = 5
    for ci, (h, w) in enumerate(headers, 1):
        cell = ws.cell(row=HDR_ROW, column=ci)
        cell.value     = h
        cell.font      = Font(name="Arial", size=9, bold=True, color=C_COL_FG)
        cell.fill      = PatternFill("solid", fgColor=C_COL_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = bdr()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[HDR_ROW].height = 30

    # ── Data rows ──
    DATA_START = HDR_ROW + 1
    for i, d in enumerate(report_data, 1):
        row = DATA_START + i - 1
        bg  = C_ALT1 if i % 2 == 0 else "FFFFFF"
        emp = d['employee']

        override = d.get('admin_debit_override')
        override_val = float(override) if override is not None else "—"
        net = d.get('net_salary', Decimal('0'))

        row_data = [
            i,
            emp.employee_id,
            emp.name,
            emp.department.name if emp.department else "—",
            emp.designation,
            d.get('present_days', 0),
            d.get('absent_days', 0),
            d.get('paid_leave_days', 0),
            d.get('total_late_minutes', 0),
            d.get('total_early_exit_minutes', 0),
            d.get('total_permission_minutes', 0),
            d.get('excess_minutes', 0),
            float(d.get('monthly_salary', 0)),
            float(d.get('absent_deduction', 0)),
            float(d.get('auto_debit_amount', 0)),
            override_val,
            float(net),
        ]

        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=ci)
            cell.value  = val
            cell.border = bdr()
            cell.font   = Font(name="Arial", size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")

            absent_hi = d.get('absent_days', 0) > 0
            late_hi   = d.get('total_late_minutes', 0) > 0 or d.get('total_early_exit_minutes', 0) > 0
            excess_hi = d.get('excess_minutes', 0) > 0

            if ci == 6:
                cell.fill = PatternFill("solid", fgColor=C_PRESENT)
                cell.font = Font(name="Arial", size=9, bold=True, color="1A7A3A")
            elif ci == 7 and absent_hi:
                cell.fill = PatternFill("solid", fgColor=C_ABSENT)
                cell.font = Font(name="Arial", size=9, bold=True, color="C0392B")
            elif ci in (9, 10, 11) and late_hi:
                cell.fill = PatternFill("solid", fgColor=C_LATE)
                cell.font = Font(name="Arial", size=9, color="8B6914")
            elif ci == 12 and excess_hi:
                cell.fill = PatternFill("solid", fgColor=C_EXCESS)
                cell.font = Font(name="Arial", size=9, bold=True, color="C0392B")
            elif ci in (14, 15) and isinstance(val, float) and val > 0:
                cell.fill = PatternFill("solid", fgColor=C_ABSENT)
                cell.font = Font(name="Arial", size=9, color="C0392B")
            elif ci == 17:
                cell.fill = PatternFill("solid", fgColor=C_NET_BG)
                cell.font = Font(name="Arial", size=10, bold=True, color=C_NET_FG)
            else:
                cell.fill = PatternFill("solid", fgColor=bg)

            if ci in (3, 4, 5):
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        ws.row_dimensions[row].height = 18

    # ── Totals row ──
    tot_row = DATA_START + len(report_data)
    ws.row_dimensions[tot_row].height = 22
    ws.merge_cells(f"A{tot_row}:E{tot_row}")
    tc = ws.cell(row=tot_row, column=1)
    tc.value     = "TOTALS"
    tc.font      = Font(name="Arial", size=10, bold=True, color=C_TOT_FG)
    tc.fill      = PatternFill("solid", fgColor=C_TOT_BG)
    tc.alignment = Alignment(horizontal="center", vertical="center")

    sum_map = {6:"F", 7:"G", 8:"H", 9:"I", 10:"J", 11:"K", 12:"L",
               13:"M", 14:"N", 15:"O", 17:"Q"}
    for ci, cl in sum_map.items():
        cell = ws.cell(row=tot_row, column=ci)
        cell.value  = f"=SUM({cl}{DATA_START}:{cl}{tot_row-1})"
        cell.fill   = PatternFill("solid", fgColor=C_TOT_BG)
        cell.font   = Font(name="Arial", size=10, bold=True,
                           color=C_NET_FG if ci == 17 else C_TOT_FG)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Override total (col 16) — manual sum since it may have "—"
    ws.cell(row=tot_row, column=16).fill = PatternFill("solid", fgColor=C_TOT_BG)

    # ── Footer ──
    fr = tot_row + 2
    ws.merge_cells(f"A{fr}:Q{fr}")
    fc = ws.cell(row=fr, column=1)
    fc.value     = "Auto-generated by NWI Attendance System  |  Confidential — Internal Use Only"
    fc.font      = Font(name="Arial", size=8, italic=True, color="888888")
    fc.alignment = Alignment(horizontal="center")

    nr = fr + 1
    ws.merge_cells(f"A{nr}:Q{nr}")
    nc = ws.cell(row=nr, column=1)
    nc.value     = "Policy: Late + Early Exit + Permissions pooled. Buffer: 120 min free. Excess = (Excess ÷ Work Minutes) × Salary."
    nc.font      = Font(name="Arial", size=8, italic=True, color="555555")
    nc.alignment = Alignment(horizontal="center")

    ws.freeze_panes = f"A{HDR_ROW+1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ============================================================
# PDF — DAILY ATTENDANCE REGISTER
# ============================================================

def generate_daily_pdf(records, selected_date, summary):
    """
    Generates styled PDF for daily attendance.
    records       : AttendanceRecord queryset for the date
    selected_date : datetime.date
    summary       : dict with present/absent/late counts
    Returns       : BytesIO buffer
    """
    buf = io.BytesIO()
    PAGE = landscape(A4)
    doc = SimpleDocTemplate(
        buf, pagesize=PAGE,
        leftMargin=12*mm, rightMargin=12*mm,
        topMargin=10*mm,  bottomMargin=12*mm,
    )

    # ── Colors ──
    DARK_GREEN   = colors.HexColor("#1A3A2A")
    MID_GREEN    = colors.HexColor("#2D5A3D")
    GOLD         = colors.HexColor("#F5C842")
    LIGHT_GREEN  = colors.HexColor("#D6F5E3")
    LIGHT_RED    = colors.HexColor("#FFE0E0")
    LIGHT_YELLOW = colors.HexColor("#FFF3CD")
    LIGHT_BLUE   = colors.HexColor("#D6EAF8")
    LIGHT_GRAY   = colors.HexColor("#F4F8F5")
    WHITE        = colors.white

    def ps(name, **kw):
        return ParagraphStyle(name, **kw)

    sty_co   = ps("co",  fontName="Helvetica-Bold", fontSize=16, textColor=GOLD,        alignment=TA_LEFT, leading=20)
    sty_sub  = ps("su",  fontName="Helvetica-Bold", fontSize=11, textColor=WHITE,        alignment=TA_LEFT, leading=14)
    sty_meta = ps("me",  fontName="Helvetica",      fontSize=8,  textColor=colors.HexColor("#AAAAAA"), alignment=TA_LEFT)
    sty_sum  = ps("sm",  fontName="Helvetica",      fontSize=8,  textColor=WHITE,        alignment=TA_RIGHT)
    sty_hdr  = ps("hd",  fontName="Helvetica-Bold", fontSize=7.5,textColor=WHITE,        alignment=TA_CENTER, leading=9)
    sty_c    = ps("ce",  fontName="Helvetica",      fontSize=8,  textColor=colors.HexColor("#1A1A1A"), alignment=TA_CENTER)
    sty_cl   = ps("cel", fontName="Helvetica",      fontSize=8,  textColor=colors.HexColor("#1A1A1A"), alignment=TA_LEFT, leftIndent=3)
    sty_mono = ps("mo",  fontName="Courier",        fontSize=7.5,textColor=colors.HexColor("#1A1A1A"), alignment=TA_CENTER)
    sty_p    = ps("sp",  fontName="Helvetica-Bold", fontSize=7.5,textColor=colors.HexColor("#1A7A3A"), alignment=TA_CENTER)
    sty_a    = ps("sa",  fontName="Helvetica-Bold", fontSize=7.5,textColor=colors.HexColor("#C0392B"), alignment=TA_CENTER)
    sty_pl   = ps("spl", fontName="Helvetica-Bold", fontSize=7.5,textColor=colors.HexColor("#1A5276"), alignment=TA_CENTER)
    sty_late = ps("sl",  fontName="Helvetica-Bold", fontSize=7.5,textColor=colors.HexColor("#8B6914"), alignment=TA_CENTER)
    sty_foot = ps("ft",  fontName="Helvetica-Oblique", fontSize=7, textColor=colors.HexColor("#888888"), alignment=TA_CENTER)

    date_str    = selected_date.strftime("%A, %d %B %Y")
    is_working  = summary.get('is_working', True)
    total_emp   = summary.get('total_active', len(list(records)))
    present_cnt = summary.get('present', 0)
    absent_cnt  = summary.get('absent', 0)
    late_cnt    = summary.get('late', 0)
    pl_cnt      = summary.get('paid_leave', 0)

    story = []

    # ── Header ──
    logo = _logo_path()
    logo_cell = ""
    if logo:
        try:
            logo_cell = RLImage(logo, width=22*mm, height=22*mm)
        except Exception:
            logo_cell = ""

    hdr_data = [[
        logo_cell,
        [
            Paragraph("Neminath Wood Industry Pvt. Ltd.", sty_co),
            Paragraph("DAILY ATTENDANCE REGISTER", sty_sub),
            Paragraph(f"Date: {date_str}   |   Total Employees: {total_emp}   |   {'Working Day' if is_working else 'Holiday'}", sty_meta),
        ],
        [
            Paragraph("SUMMARY", ps("sh", fontName="Helvetica-Bold", fontSize=8, textColor=GOLD, alignment=TA_RIGHT)),
            Paragraph(
                f"Present: {present_cnt}  |  Absent: {absent_cnt}  |  Late: {late_cnt}  |  Paid Leave: {pl_cnt}",
                sty_sum
            ),
        ],
    ]]
    hdr_t = Table(hdr_data, colWidths=[26*mm, 158*mm, 79*mm])
    hdr_t.setStyle(TableStyle([
        ("BACKGROUND",   (0,0), (-1,-1), DARK_GREEN),
        ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
        ("ALIGN",        (2,0), (2,-1),  "RIGHT"),
        ("LEFTPADDING",  (0,0), (-1,-1), 4),
        ("RIGHTPADDING", (0,0), (-1,-1), 6),
        ("TOPPADDING",   (0,0), (-1,-1), 6),
        ("BOTTOMPADDING",(0,0), (-1,-1), 6),
        ("LINEBELOW",    (0,0), (-1,-1), 2, GOLD),
    ]))
    story.append(hdr_t)
    story.append(Spacer(1, 4*mm))

    # ── Attendance Table ──
    col_hdrs = [
        Paragraph("S.No",        sty_hdr),
        Paragraph("Emp ID",      sty_hdr),
        Paragraph("Employee Name",sty_hdr),
        Paragraph("Designation", sty_hdr),
        Paragraph("Department",  sty_hdr),
        Paragraph("In Time",     sty_hdr),
        Paragraph("Out Time",    sty_hdr),
        Paragraph("Status",      sty_hdr),
        Paragraph("Late\n(min)", sty_hdr),
        Paragraph("Early\n(min)",sty_hdr),
        Paragraph("Remarks",     sty_hdr),
    ]

    STATUS_MAP = {
        "present":    ("● Present",    sty_p),
        "absent":     ("● Absent",     sty_a),
        "paid_leave": ("● Paid Leave", sty_pl),
        "holiday":    ("● Holiday",    sty_late),
    }

    table_rows  = [col_hdrs]
    row_styles  = []
    total_late  = 0
    total_early = 0

    for i, rec in enumerate(records, 1):
        rn      = i + 1
        emp     = rec.employee
        stat_txt, stat_sty = STATUS_MAP.get(rec.status, ("—", sty_c))
        late    = rec.late_minutes       or 0
        early   = rec.early_exit_minutes or 0
        total_late  += late
        total_early += early

        row = [
            Paragraph(str(i), sty_c),
            Paragraph(emp.employee_id, sty_mono),
            Paragraph(emp.name, sty_cl),
            Paragraph(emp.designation, sty_cl),
            Paragraph(emp.department.name if emp.department else "—", sty_c),
            Paragraph(rec.in_time.strftime("%H:%M")  if rec.in_time  else "—", sty_mono),
            Paragraph(rec.out_time.strftime("%H:%M") if rec.out_time else "—", sty_mono),
            Paragraph(stat_txt, stat_sty),
            Paragraph(str(late)  if late  > 0 else "—", sty_late if late  > 0 else sty_c),
            Paragraph(str(early) if early > 0 else "—", sty_late if early > 0 else sty_c),
            Paragraph(rec.remarks or "", sty_cl),
        ]
        table_rows.append(row)

        if rec.status == "absent":
            bg = LIGHT_RED
        elif rec.status == "paid_leave":
            bg = LIGHT_BLUE
        elif late > 0 or early > 0:
            bg = LIGHT_YELLOW
        elif rec.status == "present":
            bg = LIGHT_GREEN if i % 2 == 1 else WHITE
        else:
            bg = LIGHT_GRAY if i % 2 == 0 else WHITE

        row_styles.append(("BACKGROUND", (0, rn), (-1, rn), bg))

    # Totals footer row
    table_rows.append([
        Paragraph("", sty_c), Paragraph("", sty_c), Paragraph("", sty_c),
        Paragraph("", sty_c), Paragraph("", sty_c), Paragraph("", sty_c),
        Paragraph("", sty_c),
        Paragraph("TOTALS", ps("tot", fontName="Helvetica-Bold", fontSize=8, textColor=WHITE, alignment=TA_CENTER)),
        Paragraph(str(total_late),  ps("tl", fontName="Helvetica-Bold", fontSize=8, textColor=GOLD, alignment=TA_CENTER)),
        Paragraph(str(total_early), ps("te", fontName="Helvetica-Bold", fontSize=8, textColor=GOLD, alignment=TA_CENTER)),
        Paragraph("", sty_c),
    ])

    col_w = [10*mm, 18*mm, 40*mm, 32*mm, 24*mm,
             17*mm, 17*mm, 22*mm, 13*mm, 13*mm, 37*mm]

    att_t = Table(table_rows, colWidths=col_w, repeatRows=1)
    base_ts = [
        ("BACKGROUND",    (0,0),  (-1,0),  MID_GREEN),
        ("VALIGN",        (0,0),  (-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0),  (-1,-1), 4),
        ("BOTTOMPADDING", (0,0),  (-1,-1), 4),
        ("LEFTPADDING",   (0,0),  (-1,-1), 3),
        ("RIGHTPADDING",  (0,0),  (-1,-1), 3),
        ("GRID",          (0,0),  (-1,-1), 0.4, MID_GREEN),
        ("LINEBELOW",     (0,0),  (-1,0),  1.5, GOLD),
        ("BACKGROUND",    (0,-1), (-1,-1), DARK_GREEN),
        ("LINEABOVE",     (0,-1), (-1,-1), 1.5, GOLD),
    ]
    att_t.setStyle(TableStyle(base_ts + row_styles))
    story.append(att_t)

    # ── Legend ──
    story.append(Spacer(1, 3*mm))
    leg_data = [[
        Paragraph("Legend:", ps("lg", fontName="Helvetica-Bold", fontSize=7.5, textColor=DARK_GREEN)),
        Paragraph("■ Present (on time)", ps("lp", fontName="Helvetica", fontSize=7.5, textColor=colors.HexColor("#1A7A3A"))),
        Paragraph("■ Late / Early Exit",  ps("ll", fontName="Helvetica", fontSize=7.5, textColor=colors.HexColor("#8B6914"))),
        Paragraph("■ Absent",             ps("la", fontName="Helvetica", fontSize=7.5, textColor=colors.HexColor("#C0392B"))),
        Paragraph("■ Paid Leave",         ps("lpl",fontName="Helvetica", fontSize=7.5, textColor=colors.HexColor("#1A5276"))),
        Paragraph("Grace In: 8:50–9:00 AM  |  Grace Out: 6:50–7:00 PM  |  Buffer: 120 min/month",
            ps("ln", fontName="Helvetica-Oblique", fontSize=7, textColor=colors.HexColor("#555555"))),
    ]]
    leg_t = Table(leg_data, colWidths=[22*mm, 36*mm, 32*mm, 22*mm, 26*mm, 105*mm])
    leg_t.setStyle(TableStyle([
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0), (-1,-1), 2),
        ("BOTTOMPADDING", (0,0), (-1,-1), 2),
    ]))
    story.append(leg_t)

    # ── Signatures ──
    story.append(Spacer(1, 5*mm))
    sig_sty  = ps("sg",  fontName="Helvetica",      fontSize=8, textColor=DARK_GREEN, alignment=TA_CENTER)
    sig_bold = ps("sgb", fontName="Helvetica-Bold", fontSize=8, textColor=DARK_GREEN, alignment=TA_CENTER)
    sig_data = [[
        [Paragraph("_______________________", sig_sty), Paragraph("Prepared By",  sig_bold)],
        [Paragraph("_______________________", sig_sty), Paragraph("Checked By",   sig_bold)],
        [Paragraph("_______________________", sig_sty), Paragraph("Authorised By",sig_bold)],
    ]]
    sig_t = Table(sig_data, colWidths=[89*mm, 89*mm, 85*mm])
    sig_t.setStyle(TableStyle([("ALIGN",(0,0),(-1,-1),"CENTER"),("VALIGN",(0,0),(-1,-1),"MIDDLE")]))
    story.append(sig_t)

    # ── Footer ──
    story.append(Spacer(1, 3*mm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=MID_GREEN))
    story.append(Spacer(1, 2*mm))
    story.append(Paragraph(
        "Neminath Wood Industry Pvt. Ltd. — Beyond Borders  |  Auto-generated by NWI Attendance System  |  Confidential",
        sty_foot
    ))

    doc.build(story)
    buf.seek(0)
    return buf